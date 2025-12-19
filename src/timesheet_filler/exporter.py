"""PDF export functionality."""

import subprocess
import tempfile
from pathlib import Path


def export_sheets_to_pdf(
    workbook_path: Path,
    output_dir: Path,
    last_name: str,
    year: str = "26",
) -> list[Path]:
    """Export each sheet from the workbook as a separate PDF.

    Tries Excel (via AppleScript on macOS) first, then LibreOffice.
    Each sheet is exported with naming: "{LastName} {BWPeriod}-{Year}.pdf"

    Args:
        workbook_path: Path to the filled Excel workbook
        output_dir: Directory to save PDFs
        last_name: Last name for file naming
        year: Year suffix for file naming (default "26")

    Returns:
        List of paths to created PDF files
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = workbook_path.resolve()
    output_dir = output_dir.resolve()

    # Try Excel first (macOS)
    if _has_excel():
        return _export_via_excel(workbook_path, output_dir, last_name, year)

    # Fall back to LibreOffice
    soffice_path = _find_libreoffice()
    if soffice_path:
        return _export_via_libreoffice(workbook_path, output_dir, last_name, year, soffice_path)

    raise RuntimeError(
        "No PDF export method available.\n"
        "Please install either:\n"
        "  - Microsoft Excel (preferred)\n"
        "  - LibreOffice: brew install --cask libreoffice"
    )


def _has_excel() -> bool:
    """Check if Excel is available on macOS."""
    return Path("/Applications/Microsoft Excel.app").exists()


def _export_via_excel(
    workbook_path: Path,
    output_dir: Path,
    last_name: str,
    year: str,
) -> list[Path]:
    """Export sheets using Excel via AppleScript.

    Exports the entire workbook as one PDF, then splits it into
    individual single-page PDFs using pypdf.
    """
    from pypdf import PdfReader, PdfWriter
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path)
    sheet_names = wb.sheetnames
    wb.close()

    # Step 1: Export entire workbook as single PDF
    combined_pdf = output_dir / "_combined.pdf"

    applescript = f'''
tell application "Microsoft Excel"
    open "{workbook_path}"
    save active workbook in "{combined_pdf}" as PDF file format
    close active workbook without saving
end tell
'''

    print(f"    Exporting workbook to PDF...")
    result = subprocess.run(
        ["osascript", "-e", applescript],
        capture_output=True,
        text=True,
        timeout=120,
    )

    if result.returncode != 0:
        print(f"    Error: {result.stderr.strip()[:200]}")
        return []

    if not combined_pdf.exists():
        print("    Failed to create combined PDF")
        return []

    # Step 2: Split into individual PDFs
    print(f"    Splitting into {len(sheet_names)} individual PDFs...")
    pdf_paths = []

    reader = PdfReader(combined_pdf)

    for i, sheet_name in enumerate(sheet_names):
        if i >= len(reader.pages):
            break

        bw_period = str(int(sheet_name.split("-")[1]))
        pdf_name = f"{last_name} {bw_period}-{year}.pdf"
        pdf_path = output_dir / pdf_name

        writer = PdfWriter()
        writer.add_page(reader.pages[i])

        with open(pdf_path, "wb") as f:
            writer.write(f)

        pdf_paths.append(pdf_path)

    # Clean up combined PDF
    combined_pdf.unlink()

    return pdf_paths


def _find_libreoffice() -> str | None:
    """Find LibreOffice executable path."""
    mac_paths = [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/opt/homebrew/bin/soffice",
        "/usr/local/bin/soffice",
    ]

    for path in mac_paths:
        if Path(path).exists():
            return path

    try:
        result = subprocess.run(["which", "soffice"], capture_output=True, text=True)
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        pass

    return None


def _export_via_libreoffice(
    workbook_path: Path,
    output_dir: Path,
    last_name: str,
    year: str,
    soffice_path: str,
) -> list[Path]:
    """Export by creating temporary single-sheet workbooks."""
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path)
    pdf_paths = []

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)

        for sheet_name in wb.sheetnames:
            bw_period = str(int(sheet_name.split("-")[1]))

            # Create workbook with only this sheet
            temp_wb = openpyxl.load_workbook(workbook_path)
            for other_sheet in temp_wb.sheetnames:
                if other_sheet != sheet_name:
                    del temp_wb[other_sheet]

            temp_xlsx = tmpdir_path / f"{sheet_name}.xlsx"
            temp_wb.save(temp_xlsx)

            # Convert to PDF
            subprocess.run(
                [
                    soffice_path,
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(tmpdir_path),
                    str(temp_xlsx),
                ],
                capture_output=True,
                check=True,
            )

            # Rename to correct naming convention
            temp_pdf = tmpdir_path / f"{sheet_name}.pdf"
            final_pdf = output_dir / f"{last_name} {bw_period}-{year}.pdf"

            if temp_pdf.exists():
                temp_pdf.rename(final_pdf)
                pdf_paths.append(final_pdf)

    return pdf_paths
